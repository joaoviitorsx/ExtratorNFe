import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

class ExportarService:
    def __init__(self, batch_size: int = 500):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill("solid", fgColor="4F81BD")
        self.header_align = Alignment(horizontal="center")
        self.batch_size = batch_size

    def gerarPlanilha(self, dados: dict, caminho_saida: str) -> str:
        wb = Workbook(write_only=True)
        if wb.worksheets:
            wb.remove(wb.worksheets[0])

        if dados.get("validas"):
            self.notasFiscais(wb, dados["validas"])

        if dados.get("erros"):
            self.arquivosProblema(wb, dados["erros"])

        if not caminho_saida.lower().endswith(".xlsx"):
            caminho_saida += ".xlsx"

        os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)
        wb.save(caminho_saida)

        return caminho_saida

    def notasFiscais(self, wb, notas):
        ws = wb.create_sheet("Notas Fiscais")

        headers = [
            "Nome Cliente", "CNPJ Cliente", "Fornecedor", "CNPJ Fornecedor", "Inscrição Estadual",
            "Número Nota", "Série", "Data Emissão", "Data Saída/Entrada", "Chave NF",
            "Número Item", "Código Produto", "EAN", "Nome Produto", "NCM", "CFOP",
            "Quantidade", "Valor Unitário", "Valor Frete", "Valor Seguro", "Valor Desconto",
            "Valor Outros", "Valor Produto", "EAN Tributado", "Unidade Tributada", "Quantidade Tributada",
            "Valor Unitário Tributado", "Indicador Total", 
            "ICMS Origem", "ICMS CST", "ICMS Mod BC", "ICMS VBC", "ICMS PICMS", "ICMS Valor",
            "IPI CST", "IPI VBC", "IPI PIPI", "IPI VIPI",
            "PIS CST", "PIS VBC", "PIS PPIS", "PIS VPIS",
            "COFINS CST", "COFINS VBC", "COFINS PCOFINS", "COFINS VCOFINS"
        ]

        ws.append(headers)

        buffer = []
        count = 0

        for nfe in notas:
            for item in nfe.itens:
                linha = [
                    nfe.destinatario_nome,
                    nfe.destinatario_cnpj,
                    nfe.emitente_nome,
                    nfe.emitente_cnpj,
                    nfe.emitente_ie,
                    nfe.numero,
                    nfe.serie,
                    nfe.data_emissao,
                    nfe.data_saida,
                    nfe.chave,
                    item.nItem,
                    item.cProd,
                    item.cEAN,
                    item.xProd,
                    item.NCM,
                    item.CFOP,
                    item.qCom,
                    item.vUnCom,
                    item.vFrete,
                    item.vSeg,
                    item.vDesc,
                    item.vOutro,
                    item.vProd,
                    item.cEANTrib,
                    item.uTrib,
                    item.qTrib,
                    item.vUnTrib,
                    item.indTot,
                    # ICMS
                    item.icms.get("orig", "-"),
                    item.icms.get("CST", item.icms.get("CSOSN", "-")),
                    item.icms.get("modBC", "-"),
                    item.icms.get("vBC", "-"),
                    item.icms.get("pICMS", "-"),
                    item.icms.get("vICMS", "-"),
                    # IPI
                    item.ipi.get("CST", "-"),
                    item.ipi.get("vBC", "-"),
                    item.ipi.get("pIPI", "-"),
                    item.ipi.get("vIPI", "-"),
                    # PIS
                    item.pis.get("CST", "-"),
                    item.pis.get("vBC", "-"),
                    item.pis.get("pPIS", "-"),
                    item.pis.get("vPIS", "-"),
                    # COFINS
                    item.cofins.get("CST", "-"),
                    item.cofins.get("vBC", "-"),
                    item.cofins.get("pCOFINS", "-"),
                    item.cofins.get("vCOFINS", "-"),
                ]
                buffer.append(linha)
                count += 1

                if count % self.batch_size == 0:
                    for row in buffer:
                        ws.append(row)
                    buffer.clear()

        if buffer:
            for row in buffer:
                ws.append(row)

    def arquivosProblema(self, wb, arquivos):
        ws = wb.create_sheet("Arquivos com Problemas")

        headers = ["Arquivo", "Caminho"]
        ws.append(headers)

        for nome, caminho in arquivos:
            ws.append([nome, caminho])

    def cabecalho(self, ws, headers):
        for idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=idx)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_align
            ws.column_dimensions[cell.column_letter].width = max(len(header) + 2, 12)
